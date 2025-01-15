import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime
import pytz
import plotly.graph_objects as go
from Data import save_risk_data, load_risk_data, current_time
from database import get_database
from custom_logging import log_action


# """
# Emojis som anv√§nds i programmet:
# üìã : Uppgift
# üìÖ :
# ü¶∫ : Riskbed√∂mning
# üë∑ : Riskanalys
# üóÇÔ∏è : √ñversikt
# üõ†Ô∏è : Tekniska Behov
# üî® : Tekniska Behov
# üéØ : M√•l
# üìÜ : Planering
# üìä : Analys
# üìà : Kostnadsanalys
# üìâ : Historisk Data
# ‚úÖ :
# ‚úîÔ∏è :
# üîÑ : Under Arbete
# ‚ùå : Saknar uppgifter
# üéâ : Alla uppgifter klara
# üí° : Tips
# ‚åö : Arbetstid
# üõë : Inget √Ñnnu
# """

# Risk severity matrix
risk_matrix = {
    (1, 1): 1, (1, 2): 2, (1, 3): 3, (1, 4): 4,
    (2, 1): 2, (2, 2): 4, (2, 3): 6, (2, 4): 8,
    (3, 1): 3, (3, 2): 6, (3, 3): 9, (3, 4): 12,
    (4, 1): 4, (4, 2): 8, (4, 3): 12, (4, 4): 16
}


def get_severity_info(severity):
    if 1 <= severity <= 2:
        return "L√•g", "green"
    elif 3 <= severity <= 7:
        return "Medel", "yellow"
    elif 8 <= severity <= 9:
        return "Medelh√∂g", "orange"
    else:
        return "H√∂g", "red"


def create_excel_file(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Assessment"

    # Add headers
    headers = ["M√•l", "Uppgift", "Riskk√§lla", "Riskbeskrivning", "Sannolikhet", "Konsekvens",
               "Allvarlighet", "√Ötg√§rd", "Ansvarig", "Kommentarer", "Datum f√∂r √•tg√§rd",
               "Datum f√∂r Uppf√∂ljning"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Define border styles
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Add data and format cells
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=item["goal"])
        ws.cell(row=row, column=2, value=item["task"])
        ws.cell(row=row, column=3, value=item["name"])
        ws.cell(row=row, column=4, value=item["description"])
        ws.cell(row=row, column=5, value=item["likelihood"])
        ws.cell(row=row, column=6, value=item["impact"])

        severity = item["severity"]
        severity_cell = ws.cell(row=row, column=7, value=severity)
        severity_cell.fill = PatternFill(start_color=get_severity_color(severity),
                                         end_color=get_severity_color(severity),
                                         fill_type="solid")

        ws.cell(row=row, column=8, value=item["action"])
        ws.cell(row=row, column=9, value=item["responsible"])
        ws.cell(row=row, column=10, value=item["comments"])
        ws.cell(row=row, column=11, value=item["action_date"])
        ws.cell(row=row, column=12, value=item["follow_up_date"])

        # Apply borders to all cells in the row
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = thin_border

    # Set column widths
    column_widths = {
        'A': 20, 'B': 20, 'C': 15, 'D': 40, 'E': 12, 'F': 12,
        'G': 12, 'H': 40, 'I': 15, 'J': 40, 'K': 15, 'L': 15
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save to BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def get_severity_color(severity):
    _, color = get_severity_info(severity)
    return {"green": "00FF00", "yellow": "FFFF00", "orange": "FFA500", "red": "FF0000"}[color]


def display_risk_matrix():
    st.subheader("Riskmatris")

    # Create the matrix data
    data = [
        [1, 2, 3, 4],
        [2, 4, 6, 8],
        [3, 6, 9, 12],
        [4, 8, 12, 16]
    ]

    dataframe = pd.DataFrame(data)

    # Create the heatmap
    figure_risk = go.Figure(data=go.Heatmap(
        z=dataframe,
        showscale=False,
        colorscale=[
            [0, "green"], [0.16, "green"],
            [0.16, "yellow"], [0.40, "yellow"],
            [0.40, "rgb(255,165,0)"], [0.60, "rgb(255,165,0)"],
            [0.60, "red"], [1.0, "red"]
        ],
        x=["1", "2", "3", "4"],
        y=["1", "2", "3", "4"],
        text=dataframe,
        texttemplate="%{text}",
        textfont={"size": 20},
        xgap=2,
        ygap=2,
    ))

    figure_risk.update_layout(
        title="Riskbed√∂mningsmatris",
        xaxis_title="Sannolikhet",
        yaxis_title="Konsekvens",
        height=400,
        width=300,
    )

    # Display the plot
    st.plotly_chart(figure_risk, use_container_width=False)


def display_severity_descriptions():
    st.subheader("Allvarlighetsgrad")
    st.write("")
    st.write("")
    st.write("")
    st.write("")

    # Create a more visual representation with colored backgrounds
    st.markdown("""
    <style>
    .risk-level {
        padding: 4px 8px;
        border-radius: 3px;
        margin: 2px;
        color: black;
    }

    .risk-low { background-color: #00FF00; }
    .risk-medium { background-color: #FFFF00; }
    .risk-high-medium { background-color: #FFA500; color: white; }
    .risk-high { background-color: #FF0000; color: white; }
    </style>
    """, unsafe_allow_html=True)

    descriptions = [
        ("L√•g", "risk-low", "Kr√§ver ingen √•tg√§rd f√∂r tillf√§llet"),
        ("", "", ""),
        ("Medel", "risk-medium", "√Ötg√§rdas inom en rimlig tidsram"),
        ("", "", ""),
        ("Medelh√∂g", "risk-high-medium", "√Ötg√§rdas snarast"),
        ("", "", ""),
        ("H√∂g", "risk-high", "Kr√§ver omedelbar √•tg√§rd")
    ]

    for level, style_class, desc in descriptions:
        st.markdown(
            f'<div><span class="risk-level {style_class}">{level}</span> - {desc}</div>',
            unsafe_allow_html=True
        )


def get_severity_html(severity_label, severity_color, severity_value):
    color_class = {
        "green": "risk-low",
        "yellow": "risk-medium",
        "orange": "risk-high-medium",
        "red": "risk-high"
    }[severity_color]

    return f"""
    <div>
        <strong>Allvarlighet:</strong> 
        <span class="risk-level {color_class}">{severity_label}</span> 
        ({severity_value})
    </div>
    """


def parse_excel_to_risks(uploaded_file):
    """Convert uploaded Excel file to risk dictionary format"""
    try:
        df = pd.read_excel(uploaded_file)
        risks = []

        for _, row in df.iterrows():
            severity = risk_matrix[(row['Sannolikhet'], row['Konsekvens'])]
            severity_label, severity_color = get_severity_info(severity)

            risk = {
                "goal": row['M√•l'],
                "task": row['Uppgift'],
                "name": row['Riskk√§lla'],
                "description": row['Riskbeskrivning'],
                "likelihood": row['Sannolikhet'],
                "impact": row['Konsekvens'],
                "severity": severity,
                "severity_label": severity_label,
                "severity_color": severity_color,
                "action": row['√Ötg√§rd'],
                "responsible": row['Ansvarig'],
                "comments": row['Kommentarer'],
                "action_date": row['Datum f√∂r √•tg√§rd'],
                "follow_up_date": row['Datum f√∂r Uppf√∂ljning']
            }
            risks.append(risk)

        return risks, None
    except Exception as e:
        return None, f"Fel vid inl√§sning av Excel-fil: {str(e)}"


def display_risk_overview(df, risks, context="main"):
    """Display risk overview with Excel export/import functionality"""
    if not risks:
        st.info("Inga risker har lagts till √§nnu.")
        
        # Add Excel upload functionality
        st.write("Du kan ladda upp en tidigare sparad Excel-fil med risker:")
        uploaded_file = st.file_uploader(
            "V√§lj Excel-fil",
            type=["xlsx"],
            key=f"risk_excel_upload_{context}"
        )

        if uploaded_file is not None:
            imported_risks, error = parse_excel_to_risks(uploaded_file)
            if error:
                st.error(error)
            else:
                if st.button("Importera risker fr√•n Excel", key=f"import_excel_{context}"):
                    st.session_state.risks = imported_risks
                    if save_risk_data(imported_risks):
                        st.success(f"{len(imported_risks)} risker har importerats!")
                        st.rerun()
                    else:
                        st.error("Fel vid sparande av importerade risker")
        return

    st.subheader("Tillagda risker")

    # Add Excel export button at the top
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("Generera Excel", key=f"generate_excel_{context}"):
            excel_file = create_excel_file(risks)
            st.download_button(
                label="Ladda ner Excel-fil",
                data=excel_file,
                file_name="risk_assessment.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_{context}"
            )

    # Get goals for filtering
    goals = df[df['Type'] == 'Goal']['Goal_Name'].unique()

    # Create a selectbox for goal filtering
    goal_filter = st.selectbox(
        "Filtrera efter m√•l",
        ["Alla m√•l"] + list(goals),
        key=f"goal_filter_{context}"
    )

    # Filter risks based on selected goal
    filtered_risks = (
        risks if goal_filter == "Alla m√•l"
        else [risk for risk in risks if risk['goal'] == goal_filter]
    )

    # Display risks in a single level of expanders
    for i, risk in enumerate(filtered_risks):
        with st.expander(
                f"M√•l: {risk['goal']} | Uppgift: {risk['task']} | Risk: {risk['name']}",
                expanded=False
        ):
            # Create two columns for better layout
            col1, col2 = st.columns([2, 1])

            with col1:
                st.write(f"**Beskrivning:** {risk['description']}")
                st.write(f"**√Ötg√§rd:** {risk['action']}")
                if risk['comments']:
                    st.write(f"**Kommentarer:** {risk['comments']}")

            with col2:
                st.markdown(
                    get_severity_html(
                        risk['severity_label'],
                        risk['severity_color'],
                        risk['severity']
                    ),
                    unsafe_allow_html=True
                )
                st.write(f"**Ansvarig:** {risk['responsible']}")
                st.write(f"**Datum f√∂r √•tg√§rd:** {risk['action_date']}")
                st.write(f"**Datum f√∂r uppf√∂ljning:** {risk['follow_up_date']}")

            # Add metrics for quick overview
            m1, m2, m3 = st.columns(3)
            with m1:
                st.metric("Sannolikhet", risk['likelihood'])
            with m2:
                st.metric("Konsekvens", risk['impact'])
            with m3:
                st.metric("Risk", risk['severity'])


def risk_assessment_app(df):
    st.title("Riskbed√∂mning")
    col1, col2 = st.columns(2)

    with col1:
        st.write(f"Riskbed√∂mning √§r en systematisk process f√∂r att identifiera, "
                 f"analysera och utv√§rdera risker i en verksamhet, arbetsmilj√∂ eller situation. ")

    with col2:
        st.write(f"Syftet √§r att bed√∂ma sannolikheten f√∂r att n√•got o√∂nskat ska intr√§ffa och vilka konsekvenser"
                 f" det kan f√•, f√∂r att sedan vidta √•tg√§rder som minskar eller eliminerar riskerna.")
    st.divider()

    st.write(f"**Riskbed√∂mning sker i samverkan mellan Chef, Skyddsombud och Medarbetare som √§r ber√∂rda. "
             f"Denna riskbed√∂mning g√∂rs p√• sedvanligt s√§tt med hj√§lp av riskmatrisen nedan.**")

    # Load risks at the start of the function
    try:
        # Force reload risks from database
        st.session_state.risks = load_risk_data()
        print(f"Loaded {len(st.session_state.risks)} risks from database")  # Debug print
    except Exception as e:
        st.error(f"Error loading risks: {str(e)}")
        st.session_state.risks = []

    # Create tabs for input and overview
    input_tab, overview_tab = st.tabs(["L√§gg till Risk", "√ñversikt"])

    with input_tab:
        # Left side - Risk Input
        st.subheader("L√§gg till ny risk")

        # Goal and Task selection
        goals = df[df['Type'] == 'Goal']['Goal_Name'].unique()
        selected_goal = st.selectbox("V√§lj M√•l", options=goals)

        tasks = df[
            (df['Type'] == 'Task') &
            (df['Goal_Name'] == selected_goal)
            ]['Task_Name'].unique()
        selected_task = st.selectbox("V√§lj Uppgift", options=tasks)

        # Create columns for the form
        col1, col2, col3 = st.columns([5, 3, 2])

        with col1:
            # Risk input fields
            risk_name = st.text_input("Riskk√§lla")
            risk_description = st.text_area("Riskbeskrivning")
            action = st.text_input("√Ötg√§rd")
            responsible = st.text_input("Ansvarig")

        with col2:

            # Risk matrix and descriptions
            display_risk_matrix()

        with col3:
            display_severity_descriptions()

        # Bottom section for additional inputs
        likelihood = st.selectbox("Sannolikhet", options=range(1, 5))
        impact = st.selectbox("Konsekvens", options=range(1, 5))

        comments = st.text_area("Kommentarer")

        col3, col4 = st.columns(2)
        with col3:
            action_date = st.date_input("Datum f√∂r √•tg√§rd")
        with col4:
            follow_up_date = st.date_input("Datum f√∂r uppf√∂ljning")

        if st.button("L√§gg till risk"):
            if all([selected_goal, selected_task, risk_name, risk_description,
                    likelihood, impact, action, responsible]):
                severity = risk_matrix[(likelihood, impact)]
                severity_label, severity_color = get_severity_info(severity)

                new_risk = {
                    "goal": selected_goal,
                    "task": selected_task,
                    "name": risk_name,
                    "description": risk_description,
                    "likelihood": likelihood,
                    "impact": impact,
                    "severity": severity,
                    "severity_label": severity_label,
                    "severity_color": severity_color,
                    "action": action,
                    "responsible": responsible,
                    "comments": comments,
                    "action_date": action_date.strftime("%Y-%m-%d"),
                    "follow_up_date": follow_up_date.strftime("%Y-%m-%d")
                }

                st.session_state.risks.append(new_risk)
                if save_risk_data(st.session_state.risks):
                    st.success("Risk sparad i databasen!")
                    log_action("add_risk", f"{st.session_state.username} la till ny risk, {risk_name} \nf√∂r {selected_task}", "Planering/Riskbed√∂mning")
                else:
                    st.error("Fel vid sparande av risk")
            else:
                st.error("V√§nligen fyll i alla obligatoriska f√§lt")

        # Add debug information
        st.write(f"Currently loaded risks: {len(st.session_state.risks)}")

    with overview_tab:
        if not st.session_state.risks:
            st.info("Inga risker har lagts till √§nnu.")
        else:
            st.write(f"Visar {len(st.session_state.risks)} risker")
        
        display_risk_overview(df, st.session_state.risks, context="risk_app")


def create_risk_analysis(risks):
    """Create analysis graphs for risks"""
    if not risks:
        st.info("Inga risker att analysera √§nnu.")
        return

    # Convert risks to DataFrame for easier analysis
    risk_df = pd.DataFrame(risks)

    col1, col2 = st.columns([1, 2])

    with col1:
        # Risk severity distribution
        severity_counts = risk_df['severity_label'].value_counts()
        color_map = {
            'L√•g': '#00FF00',  # Green
            'Medel': '#FFFF00',  # Yellow
            'Medelh√∂g': '#FFA500',  # Orange
            'H√∂g': '#FF0000'  # Red
        }
        colors = [color_map[severity] for severity in severity_counts.index]

        fig_severity = go.Figure(data=[
            go.Bar(
                x=severity_counts.index,
                y=severity_counts.values,
                marker_color=colors,  # List of colors
                text=severity_counts.values,
                textposition='auto',
            )
        ])

        fig_severity.update_layout(
            title="F√∂rdelning av Riskallvarlighet",
            xaxis_title="Allvarlighetsgrad",
            yaxis_title="Antal",
            showlegend=False
        )

        st.plotly_chart(fig_severity, use_container_width=True)

    with col2:
        # Risks per goal - Bar chart with range slider
        goal_counts = risk_df['goal'].value_counts()
        fig_goals = go.Figure()
        
        # Add the bar chart
        fig_goals.add_trace(go.Bar(
            x=goal_counts.index,
            y=goal_counts.values,
            text=goal_counts.values,  # Add text labels
            textposition='auto',      # Automatically position labels
        ))
        
        fig_goals.update_layout(
            title="Risker per M√•l",
            xaxis_title="M√•l",
            yaxis_title="Antal Risker",
            showlegend=False,
            # Add range slider
            xaxis=dict(
                rangeslider=dict(visible=True),
                type='category',  # This ensures proper spacing for categorical data
                tickangle=-45    # Rotate labels
            ),
            # Adjust margins to accommodate rotated labels
            margin=dict(b=100),
            # Add buttons for zoom options
            updatemenus=[dict(
                type="buttons",
                showactive=False,
                # buttons=[
                #     dict(label="Reset View",
                #          method="relayout",
                #          args=[{"xaxis.range": [None, None]}])
                # ],
                x=0.05,  # Position of reset button
                y=1.15   # Position of reset button
            )]
        )
        
        st.plotly_chart(fig_goals, use_container_width=True)

    # Risk Matrix Heatmap showing number of risks at each position
    data = [
        [1, 2, 3, 4],
        [2, 4, 6, 8],
        [3, 6, 9, 12],
        [4, 8, 12, 16]
    ]
    base_matrix = pd.DataFrame(data)

    # Create risk count matrix with same shape as base matrix
    risk_counts = pd.DataFrame(0, index=range(4), columns=range(4))
    for _, risk in risk_df.iterrows():
        i = risk['likelihood'] - 1  # Adjust for 0-based indexing
        j = risk['impact'] - 1
        risk_counts.iloc[i, j] += 1

    fig_matrix = go.Figure(data=[
        # Base risk matrix
        go.Heatmap(
            z=base_matrix,
            showscale=False,
            colorscale=[
                [0, "green"], [0.16, "green"],
                [0.16, "yellow"], [0.40, "yellow"],
                [0.40, "rgb(255,165,0)"], [0.60, "rgb(255,165,0)"],
                [0.60, "red"], [1.0, "red"]
            ],
            x=["1", "2", "3", "4"],
            y=["1", "2", "3", "4"],
            text=risk_counts,  # Use risk counts for text
            texttemplate="%{text}",
            textfont={"size": 20},
            xgap=2,
            ygap=2,
        )
    ])

    fig_matrix.update_layout(
        title="Riskmatris - Antal Risker per Position",
        xaxis_title="Sannolikhet",
        yaxis_title="Konsekvens",
        height=400
    )

    st.plotly_chart(fig_matrix, use_container_width=True)

    # Summary statistics
    st.subheader("Sammanfattning")

    # Calculate most dangerous goal and task
    goal_risks = risk_df.groupby('goal').agg({
        'severity': ['mean', 'count']
    }).reset_index()
    goal_risks.columns = ['goal', 'avg_severity', 'risk_count']
    most_dangerous_goal = goal_risks.loc[goal_risks['avg_severity'].idxmax()]

    task_risks = risk_df.groupby(['goal', 'task']).agg({
        'severity': ['mean', 'count']
    }).reset_index()
    task_risks.columns = ['goal', 'task', 'avg_severity', 'risk_count']
    most_dangerous_task = task_risks.loc[task_risks['avg_severity'].idxmax()]

    # Display metrics in two rows
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Totalt antal risker", len(risks))
    with col2:
        st.metric(
            "Genomsnittlig allvarlighet",
            f"{risk_df['severity'].mean():.1f}"
        )
    with col3:
        st.metric(
            "H√∂gsta allvarlighet",
            f"{risk_df['severity'].max():.0f}"
        )
    with col4:
        high_risks = len(risk_df[risk_df['severity_label'] == 'H√∂g'])
        st.metric("Antal h√∂ga risker", high_risks)

    # Second row of metrics
    col5, col6 = st.columns(2)

    with col5:
        st.metric(
            "M√•l med h√∂gst risk",
            most_dangerous_goal['goal'],
            f"Medelv√§rde: {most_dangerous_goal['avg_severity']:.1f} ({most_dangerous_goal['risk_count']} risker)"
        )

    with col6:
        st.metric(
            "Uppgift med h√∂gst risk",
            most_dangerous_task['task'],
            f"Medelv√§rde: {most_dangerous_task['avg_severity']:.1f} ({most_dangerous_task['risk_count']} risker)"
        )

    # Timeline of risk actions
    risk_df['action_date'] = pd.to_datetime(risk_df['action_date'])
    timeline_data = risk_df.sort_values('action_date')

    fig_timeline = go.Figure(data=[
        go.Scatter(
            x=timeline_data['action_date'],
            y=timeline_data['severity'],
            mode='markers',
            marker=dict(
                size=12,
                color=timeline_data['severity'],
                colorscale=[
                    [0, "green"], [0.16, "green"],
                    [0.16, "yellow"], [0.40, "yellow"],
                    [0.40, "orange"], [0.60, "orange"],
                    [0.60, "red"], [1.0, "red"]
                ],
                showscale=True,
                colorbar=dict(title="Allvarlighet")
            ),
            text=timeline_data.apply(
                lambda x: f"M√•l: {x['goal']}<br>Risk: {x['name']}<br>Allvarlighet: {x['severity']}",
                axis=1
            ),
            hoverinfo='text'
        )
    ])

    fig_timeline.update_layout(
        title="Tidslinje f√∂r Risk√•tg√§rder",
        xaxis_title="Datum f√∂r √•tg√§rd",
        yaxis_title="Allvarlighet",
        height=400
    )

    st.plotly_chart(fig_timeline, use_container_width=True)
